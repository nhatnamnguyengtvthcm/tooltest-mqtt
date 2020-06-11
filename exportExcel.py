from openpyxl import Workbook
import server
def export_excel():
    workbook = Workbook()
    sheet = workbook.active


    # for server.receive in server.respond:
    #     print(server.receive['data'])
    #     for column, (key, value) in enumerate(server.receive['data'].items(), start=2):
    #         sheet [f"A{column}"] = key
    #         sheet [f"B{column}"] = value
    row=1
    for server.receive in server.respond:
        for (key,values) in enumerate(server.receive['data'].items()):
        # Put the key in the first column for each key in the dictionary
            sheet.cell(row=row, column=1, value=key)
            column= 2
            for element in values:
                # Put the element in each adjacent column for each element in the tuple
                sheet.cell(row=row, column=column, value=element)
                column += 1
            row += 1

    workbook.save("report.xlsx")
    print("Created excel file!")